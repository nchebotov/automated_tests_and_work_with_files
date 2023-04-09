# encoding: utf-8

import os
import csv
import allure
from io import TextIOWrapper
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from automated_tests_and_work_with_files.utils import file_management_utility


@allure.title('Archive size test and number of files in it')
def test_count_files_in_archive_and_size(open_archive):
    size_zip = file_management_utility.size_file_zip()
    assert size_zip == 1072000, f"Неверный размер архива, {size_zip} не равно 1072000"
    assert len(open_archive.filelist) == 5, f"Неверное кол-во файлов в архиве, {len(open_archive.filelist)} не равно 5"


@allure.title('String occurrence test in {csv} file')
def test_read_and_content_to_csv():
    with allure.step('Precondition'):
        path_for_resourses = file_management_utility.path_()[1]
        zip_file = ZipFile(os.path.join(path_for_resourses, "example.zip"))
    with allure.step('Open {csv} file'):
        with zip_file.open('SampleCSVFile_11kb.csv') as file:
            text = csv.reader(TextIOWrapper(file))
            with allure.step('Line-by-line writing {csv} file to list'):
                list_csv = []
                for row in text:
                    str_ = ','.join(row).replace(',', ',', 11)
                    list_csv.append(str_)
    with allure.step('Asserting for occerrence text in {csv} file'):
        assert "11,Xerox 1980,Neola Schneider,807,-166.85,4.28,6.18,Nunavut,Paper,0.4" in list_csv
    with allure.step('Postcondition'):
        file.close()


@allure.title('Text occurrence test in {xlsx} file')
def test_read_and_content_to_xlsx():
    path_for_resourses = file_management_utility.path_()[1]
    zip_file = ZipFile(os.path.join(path_for_resourses, "example.zip"))
    with zip_file.open("file_example_XLSX_10.xlsx") as file_xlsx:
        text = load_workbook(file_xlsx)
        sheet = text.active
        assert sheet.cell(row=8, column=7).value == "15/10/2017", f"Ожидаем {'15/10/2017'}," \
                                                                  f"фактически имеем {sheet.cell(row=8, column=7).value}"
    file_xlsx.close()


@allure.title('Text occurrence test in {pdf} file')
def test_read_and_content_to_pdf():
    path_for_resourses = file_management_utility.path_()[1]
    zip_file = ZipFile(os.path.join(path_for_resourses, 'example.zip'))
    with zip_file as file_zip:
        file_pdf = file_zip.extract("file-example_PDF_1MB.pdf")
        reader = PdfReader(file_pdf)
        try:
            page = reader.pages[0]
            assert "Morbi viverra semper lorem nec molestie" in page.extract_text(), \
                f"Фраза {'Morbi viverra semper lorem nec molestie'} отсутствует в файле {'file-example_PDF_1MB.pdf'}"
        finally:
            # base_path = file_management_utility.path_()[2]
            # os.remove(os.path.join(base_path, 'file-example_PDF_1MB.pdf'))
            file_zip.close()
